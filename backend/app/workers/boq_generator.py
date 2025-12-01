"""
BOQ generation utilities.

This module takes the rich `cad_data` structure produced by
`ComprehensiveCADExtractor` and turns it into:

- A normalized BOQ data dictionary (`BOQGenerator`)
- A formatted Excel file (`ExcelBOQWriter`)

It is intentionally conservative and generic so it can work with
different drawings without needing project‑specific rules.
"""

from __future__ import annotations

from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


BOQItemDict = Dict[str, Any]


@dataclass
class BOQItem:
    """Represents a single BOQ line item."""

    sl_no: int
    description: str
    category: str
    quantity: float
    unit: str = "Nos"
    confidence: str = "medium"
    source: Optional[str] = None  # e.g. BLOCK name / layer

    def to_dict(self) -> BOQItemDict:
        return asdict(self)


class BOQGenerator:
    """
    Lightweight BOQ generator that groups entities into countable items.

    The implementation is deliberately simple and robust:
    - Counts block INSERTs grouped by block name
    - Uses basic keyword heuristics to assign categories
    - Produces a structure compatible with the pipeline expectations:
        {
          "items": [ ... ],
          "statistics": {
             "total_items": int,
             "categories": int,
             "high_confidence": int
          }
        }
    """

    def __init__(self, cad_data: Dict[str, Any]):
        self.cad_data = cad_data or {}
        self.entities = self.cad_data.get("entities") or {}
        # Ensure entities is always a dict, not None
        if not isinstance(self.entities, dict):
            self.entities = {}

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def generate(self) -> Dict[str, Any]:
        """
        Build a BOQ dictionary from CAD data.

        This will never raise on missing sections – at worst it will
        return an empty BOQ with zeroed statistics so Celery can still run.
        """
        items: List[BOQItem] = []
        sl_no = 1

        # 1) Block inserts → count as discrete items
        for item in self._generate_from_inserts(start_sl_no=sl_no):
            items.append(item)
            sl_no += 1

        # Future: extend with walls, rooms, hatches, etc.

        statistics = self._build_statistics(items)

        return {
            "items": [item.to_dict() for item in items],
            "statistics": statistics,
        }

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------
    def _generate_from_inserts(self, start_sl_no: int = 1) -> Iterable[BOQItem]:
        # Ensure entities is a dict
        if not isinstance(self.entities, dict):
            return []
        
        inserts = self.entities.get("inserts") or []
        if not inserts:
            return []

        # Group by block name
        counts: Dict[str, int] = {}
        for ins in inserts:
            # Skip None entries
            if ins is None:
                continue
            # Handle both dict and object-like access
            if isinstance(ins, dict):
                name = (ins.get("block_name") or "").strip()
            else:
                name = (getattr(ins, "block_name", None) or "").strip()
            if not name:
                continue
            counts[name] = counts.get(name, 0) + 1

        items: List[BOQItem] = []
        sl_no = start_sl_no
        for block_name, qty in sorted(counts.items(), key=lambda kv: kv[0].lower()):
            category, confidence = self._infer_category(block_name)
            description = block_name  # You can customize this later
            items.append(
                BOQItem(
                    sl_no=sl_no,
                    description=description,
                    category=category,
                    quantity=float(qty),
                    unit="Nos",
                    confidence=confidence,
                    source=f"BLOCK:{block_name}",
                )
            )
            sl_no += 1

        return items

    def _infer_category(self, name: str) -> (str, str):
        """
        Very simple keyword‑based categorization based on block name.
        Returns (category, confidence).
        """
        upper = name.upper()

        mapping = [
            (["DOOR", "DOR", "PORTE"], "Doors"),
            (["WINDOW", "WIN", "FENETRE"], "Windows"),
            (["CHAIR", "SEAT"], "Chairs"),
            (["TABLE", "DESK"], "Tables"),
            (["BED"], "Beds"),
            (["SOFA", "COUCH"], "Sofas"),
            (["WARDROBE", "CUPBOARD"], "Storage"),
        ]

        for keywords, category in mapping:
            if any(kw in upper for kw in keywords):
                return category, "high"

        # Fallback generic categories
        if "WALL" in upper:
            return "Walls", "medium"
        if "COLUMN" in upper or "COL" in upper:
            return "Columns", "medium"

        return "Miscellaneous", "low"

    def _build_statistics(self, items: List[BOQItem]) -> Dict[str, Any]:
        categories = {item.category for item in items}
        high_conf = sum(1 for item in items if item.confidence == "high")

        return {
            "total_items": len(items),
            "categories": len(categories),
            "high_confidence": high_conf,
        }


class ExcelBOQWriter:
    """
    Writes BOQ data to a formatted Excel workbook.

    Expects data in the shape returned by `BOQGenerator.generate()`
    plus a `project_info` dictionary with arbitrary metadata.
    """

    def __init__(self, boq_data: Dict[str, Any], project_info: Dict[str, Any]):
        self.boq_data = boq_data or {}
        self.items: List[Dict[str, Any]] = self.boq_data.get("items") or []
        self.statistics: Dict[str, Any] = self.boq_data.get("statistics") or {}
        self.project_info = project_info or {}

    def write(self, output_path: str | Path) -> str:
        """
        Create the Excel file at `output_path`.

        Returns the absolute path to the created file.
        """
        output_path = str(Path(output_path).absolute())

        wb = Workbook()
        ws = wb.active
        ws.title = "BOQ"

        self._write_header(ws)
        self._write_items(ws, start_row=8)
        self._auto_size_columns(ws)

        # Optional: a simple "Summary" sheet
        self._write_summary_sheet(wb)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    # ------------------------------------------------------------------
    # Sheet helpers
    # ------------------------------------------------------------------
    def _write_header(self, ws):
        # Project info block
        title_font = Font(bold=True, size=14)
        ws["A1"] = "Bill of Quantities"
        ws["A1"].font = title_font
        ws.merge_cells("A1:G1")
        ws["A1"].alignment = Alignment(horizontal="center")

        row = 3
        info_font = Font(bold=True)
        for key, value in self.project_info.items():
            ws[f"A{row}"] = str(key)
            ws[f"A{row}"].font = info_font
            ws[f"B{row}"] = str(value)
            row += 1

        # Column headers
        header_row = 7
        headers = [
            "Sl No",
            "Item Description",
            "Category",
            "Quantity",
            "Unit",
            "Confidence",
            "Source",
        ]
        header_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def _write_items(self, ws, start_row: int):
        row = start_row
        for item in self.items:
            # We accept both BOQItem and dict representations
            if isinstance(item, BOQItem):
                data = item.to_dict()
            else:
                data = dict(item)

            ws.cell(row=row, column=1, value=data.get("sl_no"))
            ws.cell(row=row, column=2, value=data.get("description"))
            ws.cell(row=row, column=3, value=data.get("category"))
            ws.cell(row=row, column=4, value=data.get("quantity"))
            ws.cell(row=row, column=5, value=data.get("unit"))
            ws.cell(row=row, column=6, value=data.get("confidence"))
            ws.cell(row=row, column=7, value=data.get("source"))
            row += 1

    def _auto_size_columns(self, ws):
        for column_cells in ws.columns:
            length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    value = str(cell.value) if cell.value is not None else ""
                    length = max(length, len(value))
                except Exception:
                    continue
            ws.column_dimensions[column_letter].width = min(max(length + 2, 10), 50)

    def _write_summary_sheet(self, wb: Workbook):
        ws = wb.create_sheet(title="Summary")

        ws["A1"] = "BOQ Summary"
        ws["A1"].font = Font(bold=True, size=14)

        stats = {
            "Total BOQ Items": self.statistics.get("total_items", 0),
            "Distinct Categories": self.statistics.get("categories", 0),
            "High Confidence Items": self.statistics.get("high_confidence", 0),
        }

        row = 3
        key_font = Font(bold=True)
        for key, value in stats.items():
            ws[f"A{row}"] = key
            ws[f"A{row}"].font = key_font
            ws[f"B{row}"] = value
            row += 1


