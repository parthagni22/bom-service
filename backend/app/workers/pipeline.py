"""
backend/app/workers/pipeline.py - FIXED FOR WINDOWS
Replace your existing pipeline.py with this version
"""

import os
import subprocess
import shutil
from collections import defaultdict
from openpyxl import Workbook
from .dxf_extract import read_dxf_entities
from .catalog import load_catalog_map, normalize_row
from .validators import validate_required
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def _resolve_converter_bin(converter: str) -> str:
    """
    Locate the converter executable. Prefer CONVERTER_BIN, otherwise try to
    auto-detect common installation paths so local runs don't fail with a
    placeholder path.
    """
    configured = os.getenv("CONVERTER_BIN")
    if configured and os.path.exists(configured):
        return configured

    if converter == "oda":
        program_files = os.environ.get("ProgramFiles", r"C:\Program Files")
        oda_root = os.path.join(program_files, "ODA")
        candidates = [
            os.path.join(program_files, "ODA", "ODAFileConverter.exe"),
            os.path.join(program_files, "ODAFileConverter", "ODAFileConverter.exe"),
        ]
        if os.path.isdir(oda_root):
            for root, _, files in os.walk(oda_root):
                if "ODAFileConverter.exe" in files:
                    candidates.insert(0, os.path.join(root, "ODAFileConverter.exe"))
                    break

        for candidate in candidates:
            if os.path.exists(candidate):
                return candidate

        raise FileNotFoundError(
            "ODAFileConverter.exe not found. Set CONVERTER_BIN to the full path "
            "of your ODA FileConverter installation."
        )

    bin_name = configured or "dwg2dxf"
    resolved = shutil.which(bin_name)
    if resolved:
        return resolved

    raise FileNotFoundError(
        f"Converter '{bin_name}' not found in PATH. Set CONVERTER_BIN accordingly."
    )


def run_pipeline(job_id: str, in_path: str, base_dir: str) -> dict:
    """
    Enhanced pipeline with proper Windows path handling
    
    job_id: UUID
    in_path: path to uploaded DWG
    base_dir: backend/../data/jobs
    """
    logger.info(f"[{job_id}] Starting pipeline for: {in_path}")
    
    job_root = os.path.join(base_dir, job_id)
    in_dir = os.path.join(job_root, "in")
    out_dir = os.path.join(job_root, "out")
    tmp_dir = os.path.join(job_root, "tmp")
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(tmp_dir, exist_ok=True)

    # 1) Convert DWG -> DXF
    converter = os.getenv("CONVERTER", "oda")
    dxf_path = os.path.join(tmp_dir, "layout.dxf")

    try:
        if converter == "oda":
            # ODA FileConverter with proper Windows path handling
            oda_bin = _resolve_converter_bin(converter)
            
            # CRITICAL FIX: Normalize paths for Windows
            oda_bin = os.path.normpath(oda_bin)
            in_dir_norm = os.path.normpath(in_dir)
            tmp_dir_norm = os.path.normpath(tmp_dir)
            
            # Verify ODA exists
            if not os.path.exists(oda_bin):
                raise FileNotFoundError(f"ODA FileConverter not found at: {oda_bin}")
            
            logger.info(f"[{job_id}] Using ODA FileConverter: {oda_bin}")
            
            # Build command - ODA syntax: 
            # ODAFileConverter <inDir> <outDir> <outVer> <outType> <recurse> <audit>
            cmd = [
                oda_bin,
                in_dir_norm,
                tmp_dir_norm,
                os.getenv("DXF_VERSION", "ACAD2018"),
                "DXF",
                "0",  # Don't recurse subdirectories
                "1"   # Audit and recover
            ]
            
            logger.warning(f"Running command:\n")
            logger.warning(f"{cmd}")
            
            # Run with shell=True on Windows to handle spaces in paths
            result = subprocess.run(
                cmd,
                shell=False,  # Keep False but paths are now normalized
                capture_output=True,
                text=True,
                timeout=300,
                cwd=tmp_dir  # Set working directory
            )
            
            if result.returncode != 0:
                logger.error(f"ODA stderr: {result.stderr}")
                logger.error(f"ODA stdout: {result.stdout}")
                raise RuntimeError(f"ODA conversion failed with code {result.returncode}")
            
            logger.info(f"ODA stdout: {result.stdout}")
            
            # Find produced DXF (ODA preserves filename)
            candidates = [os.path.join(tmp_dir, f) for f in os.listdir(tmp_dir) if f.lower().endswith(".dxf")]
            if not candidates:
                raise RuntimeError("DXF not produced by ODA converter. Check if DWG file is valid.")
            dxf_path = candidates[0]
            logger.info(f"[{job_id}] DXF created: {dxf_path}")
            
        else:
            # LibreDWG fallback
            lib_bin = _resolve_converter_bin(converter)
            logger.info(f"[{job_id}] Using LibreDWG: {lib_bin}")
            
            cmd = [lib_bin, in_path, "-o", dxf_path]
            subprocess.check_call(cmd, timeout=300)
            
            if not os.path.exists(dxf_path):
                raise RuntimeError("DXF not produced by LibreDWG")
    
    except subprocess.TimeoutExpired:
        raise RuntimeError(f"Conversion timed out after 300 seconds")
    except FileNotFoundError as e:
        logger.error(f"Converter binary not found: {e}")
        raise RuntimeError(f"Converter not found. Please check CONVERTER_BIN in .env file. Error: {e}")
    except Exception as e:
        logger.error(f"Conversion failed: {e}")
        raise

    # Verify DXF was created
    if not os.path.exists(dxf_path):
        raise RuntimeError(f"DXF file was not created at: {dxf_path}")
    
    logger.info(f"[{job_id}] DXF size: {os.path.getsize(dxf_path)} bytes")

    # 2) Parse DXF -> rows
    logger.info(f"[{job_id}] Parsing DXF entities...")
    try:
        entities = list(read_dxf_entities(dxf_path))
        logger.info(f"[{job_id}] Found {len(entities)} entities")
    except Exception as e:
        logger.error(f"DXF parsing failed: {e}")
        raise RuntimeError(f"Failed to parse DXF: {e}")

    if len(entities) == 0:
        logger.warning(f"[{job_id}] No entities found in DXF. Check if drawing has blocks/content.")

    # 3) Normalize via catalog
    catalog_path = os.path.join(os.path.dirname(__file__), "..", "resources", "catalog_map.csv")
    catalog_path = os.path.abspath(catalog_path)
    catalog = load_catalog_map(catalog_path)
    rows = [normalize_row(e, catalog) for e in entities]

    # 4) Validate
    exceptions = validate_required(rows)
    logger.info(f"[{job_id}] Found {len(exceptions)} validation exceptions")

    # 5) Aggregate -> BOQ
    agg = defaultdict(lambda: {"qty": 0, "layers": set()})
    for r in rows:
        key = (r["item_code"] or r["block"], r["desc"], r["size"], r["material"], r["uom"])
        agg[key]["qty"] += 1
        agg[key]["layers"].add(r["layer"])

    logger.info(f"[{job_id}] Aggregated into {len(agg)} unique items")

    # 6) Write Excel
    xlsx = os.path.join(out_dir, "BOQ_Output.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ_Master"
    
    # Headers
    ws.append(["Item Code", "Description", "Size/Spec", "Material", "UOM", "Qty", "Layers"])
    
    # Data rows
    for (code, desc, size, material, uom), info in agg.items():
        ws.append([code, desc, size, material, uom, info["qty"], ", ".join(sorted(info["layers"]))])

    # Room-wise sheet (optional)
    room_ws = wb.create_sheet("Room_Wise")
    room_ws.append(["Room/Zone", "Item Code", "Description", "Qty"])
    room_totals = defaultdict(int)
    for r in rows:
        room = r["room"] or ""
        if room:
            room_totals[(room, r["item_code"] or r["block"], r["desc"])] += 1
    for (room, code, desc), qty in room_totals.items():
        room_ws.append([room, code, desc, qty])

    # Exceptions sheet
    ex = wb.create_sheet("Unmapped_Exceptions")
    ex.append(["Block", "Layer", "Item Code", "Desc", "Size", "Material", "Room", "Issue"])
    for bad in exceptions:
        ex.append([bad["block"], bad["layer"], bad["item_code"], bad["desc"], 
                   bad["size"], bad["material"], bad["room"], bad["issue"]])

    wb.save(xlsx)
    logger.info(f"[{job_id}] Excel saved: {xlsx}")

    return {
        "job_id": job_id,
        "status": "success",
        "output": xlsx,
        "items_parsed": len(rows),
        "unique_items": len(agg),
        "exceptions": len(exceptions)
    }